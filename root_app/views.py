from django.shortcuts import render, redirect
from django.contrib.auth.decorators import user_passes_test, login_required, permission_required
from root_app.models import User
from django.http import JsonResponse
from e_voting_13.random_password import generate_password
from django.contrib.auth.models import Group, Permission
from django.db.models import Q
from root_app.models import (Programme, YearClass, 
                             Hall, PollingStation, 
                             Election, ElectoralCommissioner, 
                             AllowedPollingStation, Programme,
                             Hall, PollingStation, ElectorateProfile, 
                             Position, Aspirant, VerifiedElectorate, 
                             CurrentElection, Vote, InstitutionInfo)
from openpyxl import load_workbook
from django.urls import reverse
from .emails import send_password_resetlink


def is_superuser(user):

    if user.is_authenticated:

        try:
            user.user_type == 'admin'
            return 'admin'
        
        except User.DoesNotExist:
            return False

    else:
        return False


@login_required(login_url='authentication_app:login')
def dashboard(request):

    electorates = ElectorateProfile.objects.all().count()

    users = User.objects.all().order_by('time_stamp')[:5]
    
    filtered_users = []

    for data in users:
        if data.user_type == 'superuser' or data.user_type == 'staff':

            filtered_users.append(
                {
                    'username':data.username,
                    'full_name':data.full_name,
                    'date_added':data.time_stamp,
                    'last_login':data.last_login,
                    'user_type':data.user_type,
                    'email':data.email,
                    'id':data.id,
                }
            )

    elections = Election.objects.all()     

    context = {
        'electorates_total':electorates,
        'elections':elections
    }

    context.update(
        {'users':filtered_users}
    )

    if request.user.has_perm('root_app.add_election') and request.user.user_type != 'superuser' and not request.user.has_perm('root_app.can_assign_commnissioner_role'):
        #Counting Elections
        total_election = request.user.electoralcommissioner_set.count()

        #Counting Aspirants according to current election
        commisioner = ElectoralCommissioner.objects.filter(user=request.user).first() # Getting the current election by getting the electoral commissioner

        total = Aspirant.objects.filter(election=commisioner.election).count() #Filtering  and counting Aspirants

        context.update(
            {'total_elections':total_election, 'total_aspirants':total}
        )

    else:
        context.update(
            {'total_aspirants':Aspirant.objects.count(), 'total_elections':Election.objects.count()}
        )
    

    return render(request, 'root_app/dashboard.html', context)



def getting_totalaspirants(request):

    election_id = request.POST['election']

    if election_id != '':

        election = Election.objects.get(id=election_id)

        total = Aspirant.objects.filter(election=election).count() #Filtering  and counting Aspirants

        return JsonResponse({'total_aspirants':total})

    else:
    
        if request.user.has_perm('root_app.add_election') and request.user.user_type != 'superuser' and not request.user.has_perm('root_app.can_assign_commnissioner_role'):

            total = 0

            commissions = ElectoralCommissioner.objects.filter(user=request.user)

            for data in commissions:
                total += Aspirant.objects.filter(election=data.election).count()
            
            return JsonResponse({'total_aspirants':total})
            
        else:

            return JsonResponse({'total_aspirants': Aspirant.objects.count()})



@login_required(login_url='athentication_app:login')
def permissible_page(request):

    return render(request, 'root_app/perm_page.html')



@login_required(login_url='authentication_app:login')
@permission_required('authentication_app.add_user', login_url='root_app:permissible_page')
def add_user(request):

    if request.method == 'POST':
        
        username = request.POST['username']
        email = request.POST['email']
        first_name = request.POST['first_name']
        last_name = request.POST['last_name']
        other_name = request.POST['other_name']
        phone = request.POST['phone_number']
        user_type = request.POST['user_type']

        if username == '':
            return JsonResponse({'code':400, 'message':'Username required'})
        
        elif User.objects.filter(username=username).exists():
            return JsonResponse({'code':400, 'message':'Email already exist'})
        
        elif User.objects.filter(email=email).exists():
            return JsonResponse({'code':400, 'message':'Email already exist'})
        
        elif User.objects.filter(phone_number=phone).exists():
            return JsonResponse({'code':400, 'message':'Phone number already exist'})
        
        elif phone == '':
            return JsonResponse({'code':400, 'message':'Phone number required'})
        
        elif email == '':
            return JsonResponse({'code':400, 'message':'Email required'})
        
        elif first_name == '':
            return JsonResponse({'code':400, 'message':'Email required'})
        
        elif last_name == '':
            return JsonResponse({'code':400, 'message':'Last name required'})
        
        elif user_type == '':
            return JsonResponse({'code':400, 'message':'Select user type'})

        if user_type == 'superuser':
            user = User.objects.create_superuser(
                username=username,
                email=email,
                first_name=first_name,
                last_name=last_name,
                other_name=other_name,
                password=generate_password(10),
                phone_number=phone
            )

            try:
                request.POST['status']
                pass
            except KeyError:
                user.active = False
                user.save()

            return JsonResponse({'code':200, 'message':'User added successfully'})

        elif user_type == 'staff':
            user = User.objects.create_staffuser(
                username=username,
                email=email,
                first_name=first_name,
                last_name=last_name,
                other_name=other_name,
                password=generate_password(10),
                phone_number=phone
            )
            
            try:
                request.POST['status']
                pass
            except KeyError:
                user.active = False
                user.save()

            return JsonResponse({'code':200, 'message':'User added successfully'})
        
        else:
            user = User.objects.create_user(
                username=username,
                email=email,
                first_name=first_name,
                last_name=last_name,
                other_name=other_name,
                password=generate_password(10),
                phone_number=phone

            )
            
            try:
                request.POST['status']
                pass
            except KeyError:
                user.active = False
                user.save()

            return JsonResponse({'code':200, 'message':'User added successfully'})

    else:

        return render(request, 'root_app/user.html')


@login_required(login_url='authentication_app:login')
@permission_required('authentication_app.view_users', login_url='root_app:permissible_page')
def users(request):

    users = User.objects.all()

    context = {
        'users': users
    }

    return render(request, 'root_app/users.html', context)
    

@login_required(login_url='authentication_app:login')
@permission_required('authentication_app.change_user', login_url='root_app:permissible_page')
def edit_user(request, user_id):

    if request.method == 'POST':
        username = request.POST['username']
        email = request.POST['email']
        first_name = request.POST['first_name']
        last_name = request.POST['last_name']
        other_name = request.POST['other_name']
        phone = request.POST['phone_number']
        user_type = request.POST['user_type']

        user = User.objects.get(id=user_id)

        if username == '':
            return JsonResponse({'code':400, 'message':'Username required'})
        
        elif User.objects.filter(username=username).exists() and user.username != username:
            return JsonResponse({'code':400, 'message':'Email already exist'})
        
        elif User.objects.filter(email=email).exists() and user.email != email:
            return JsonResponse({'code':400, 'message':'Email already exist'})
        
        elif User.objects.filter(phone_number=phone).exists() and user.phone_number != phone:
            return JsonResponse({'code':400, 'message':'Phone number already exist'})
        
        elif phone == '':
            return JsonResponse({'code':400, 'message':'Phone number required'})
        
        elif email == '':
            return JsonResponse({'code':400, 'message':'Email required'})
        
        elif first_name == '':
            return JsonResponse({'code':400, 'message':'Email required'})
        
        elif last_name == '':
            return JsonResponse({'code':400, 'message':'Last name required'})
        
        elif user_type == '':
            return JsonResponse({'code':400, 'message':'Select user type'})

        if user_type == 'superuser':
            user.username=username
            user.email=email
            user.first_name=first_name
            user.last_name=last_name
            user.other_name=other_name
            user.phone_number=phone
            user.superuser=True
            user.staff=True

            try:
                request.POST['status']
                user.active = True
            except KeyError:
                user.active = False
            
            user.save()

            return JsonResponse({'code':200, 'message':'User updated successfully'})

        elif user_type == 'staff':
            user.username=username
            user.email=email
            user.first_name=first_name
            user.last_name=last_name
            user.other_name=other_name
            user.phone_number=phone
            user.superuser=False
            user.staff=True
            
            try:
                request.POST['status']
                user.active = True
            except KeyError:
                user.active = False
            
            user.save()

            return JsonResponse({'code':200, 'message':'User updated successfully'})
        
        else:
            user.username=username
            user.email=email
            user.first_name=first_name
            user.last_name=last_name
            user.other_name=other_name
            user.phone_number=phone
            user.superuser=False
            user.staff=False
            
            try:
                request.POST['status']
                user.active = True
            except KeyError:
                user.active = False
            
            user.save()

            return JsonResponse({'code':200, 'message':'User updated successfully'})
    else:

        try:

            user = User.objects.get(id=user_id)
            
            context = {
                'user':user
            }

            return render(request, 'root_app/edit_user.html', context)


        except User.DoesNotExist:
            return render(request, 'root_app/error-404.html')
        
        except ValueError:
            return render(request, 'root_app/error-404.html')


@login_required(login_url='authentication_app:login')
@permission_required('authentication_app.view_user', login_url='root_app:permissible_page')
def view_user(request, user_id):

    if request.method =='POST':
        
        group_ids = request.POST.getlist('groups')

        user = User.objects.get(id=user_id)

        if not request.user.has_perm('authentication_app.assign_group'):
            return JsonResponse({'code':400, 'message':'You are not permittable'})

        user.groups.clear()

        for data in group_ids:
            user.groups.add(Group.objects.get(id=data))
        
        return JsonResponse({'code':200, 'message':'Groups assigned successfully'})

    else:
        try:

            permissions = Permission.objects.all()
            
            user = User.objects.get(id=user_id)

            groups = []

            gps = Group.objects.all()

            for data in gps:
                
                if user.groups.filter(id=int(data.id)).exists():
                    continue
                else:
                    groups.append(
                        {'id':data.id, 'name':data.name}
                    )

            selected_groups = user.groups.all()
                    
            context = {
                'data':user,
                'groups':groups,
                'seleted_groups':selected_groups,
                'permissions':permissions,
            }

            return render(request, 'root_app/view_user.html', context)
        
        except User.DoesNotExist:
            return render(request, 'root_app/error-404.html')
        
        except ValueError:
            return render(request, 'root_app/error-404.html')



@login_required(login_url='authentication_app:login')
@permission_required('authentication_app.delete_user', login_url='root_app:permissible_page')
def delete_user(request):

    try:

        user_id = request.POST['user_id']

        User.objects.get(id=user_id).delete()

        return JsonResponse({'code':200, 'message':'User has been deleted successfully'})
    
    except User.DoesNotExist:
        return JsonResponse({'code':400, 'message':'Invalid user id or does not exist'})
    
    except Exception:
        return JsonResponse({'code':400, 'message':'Something went wrong, try again'})


@login_required(login_url='authentication_app:login')
@permission_required({('auth.add_group'), ('auth.change_group')}, login_url='root_app:permissible_page')
def group(request, group_id=None):

    is_update = False

    if group_id is not None:
        is_update = True  

    if request.method == 'POST':
        group_name = request.POST['group']

        if group_name == '':
            return JsonResponse({'code':400, 'message':'Group name required'})

        if not is_update:
            group = Group.objects.create(
                name=group_name
            )
            group.save()
            return JsonResponse({'code':200, 'message':'Group added successfully'})
        else:
            group = Group.objects.get(id=group_id)
            group.name = group_name
            group.save()
            return JsonResponse({'code':200, 'message':'Group is has been updated successfully'})
    else:
        
        if is_update:

            try:
                group = Group.objects.get(id=group_id)

            except Group.DoesNotExist:
                return render(request, 'root_app/error-404.html')
                
            except ValueError:
                return render(request, 'root_app/error-404.html')

            context = {
                'group': group,
                'is_update':is_update,
            }
            return render(request, 'root_app/group.html', context)
        
        return render(request, 'root_app/group.html')



@login_required(login_url='authentication_app:login')
@permission_required('auth.view_group', login_url='root_app:permissible_page')
def groups(request):

    groups = Group.objects.all()

    context = {
        'groups':groups
    }

    return render(request, 'root_app/groups.html', context)


@login_required(login_url='authentication_app:login')
@permission_required('auth.delete_group', login_url='root_app:permissible_page')
def delete_group(request):

    try:

        group_id = request.POST['group_id']

        Group.objects.get(id=group_id).delete()

        return JsonResponse({'code':200, 'message':'Group has been deleted successfully'})
    
    except User.DoesNotExist:
        return JsonResponse({'code':400, 'message':'Invalid group id or does not exist'})
    
    except Exception:
        return JsonResponse({'code':400, 'message':'Something went wrong, try again'})


@login_required(login_url='authentication_app:login')
@permission_required({('auth.add_permission'), ('auth.change_permission')}, login_url='root_app:permissible_page')
def group_permissions(request, group_id):

    if request.method == 'POST':

        try:
            sel_permissions = request.POST.getlist('perm_ids[]')

            # if len(sel_permissions) <= 0:
            #     return JsonResponse({'code':400, 'message':'No permission selected'})

            group = Group.objects.get(id=group_id)

            group.permissions.clear()

            for data in sel_permissions:

                permission = Permission.objects.get(id=data)

                group.permissions.add(permission)

            return JsonResponse({'code':200, 'message':'Saved successfully'})

        except Exception as ex:
            return JsonResponse({'code':400, 'message':'Something went wrong try again'})
        
    else:

        try:

            group = Group.objects.get(id=group_id)

            request.session['group_id'] = group.id

        except Group.DoesNotExist:
            return render(request, 'root_app/error-404.html')
        
        except ValueError:
            return render(request, 'root_app/error-404.html')

        permissions = []

        perms = Permission.objects.all()

        for data in perms:

            if group.permissions.filter(id=data.id).exists():
                continue
            else:
                permissions.append(
                    {
                        'id':data.id,
                        'name':data.name,
                        'codename':data.codename
                    }
                )
    
        group_permissions = group.permissions.all()

        context = {
            'permissions':permissions,
            'group':group,
            'group_permissions':group_permissions
        }

        return render(request, 'root_app/group_permissions.html', context)



@login_required(login_url='authentication_app:login')
def search_permissions(request):

    search_item = request.POST['search_item']

    perms = Permission.objects.filter(
        Q(name__icontains=search_item) | Q(codename__icontains=search_item)
    )

    permission = []

    group = Group.objects.get(id=request.session['group_id'])

    for data in perms:
        if group.permissions.filter(id=data.id).exists():
            continue
        else:
            permission.append(
                {
                    'id':data.id,
                    'name':data.name,
                    'codename':data.codename
                }
            )

    return JsonResponse({'permissions':permission})


@login_required(login_url='authentication_app:login')
@permission_required('root_app.add_programme', login_url='root_app:permissible_page')
def programme(request, programme_id=None):

    is_update = False

    if programme_id is not None:
        is_update = True

    if request.method == 'POST':
        programme_name = request.POST['programme_name']

        if programme_name == '':
            return JsonResponse({'code':400, 'message':'Programme name required'})
        
        if not is_update:

            programme = Programme.objects.create(
                programme_name=programme_name
            )
            programme.save()

            return JsonResponse({'code':200, 'message':'Programme added successfully'})

        else:

            programme = Programme.objects.get(id=programme_id)
            programme.programme_name = programme_name
            programme.save()

            return JsonResponse({'code':200, 'message':'Programme updated successfully'})

    else:

        if is_update:

            if not request.user.has_perm('root_app.change_programme'):
                return redirect('root_app:permissible_page')
        
            else:

                try:

                    programme = Programme.objects.get(id=programme_id)
                
                except Programme.DoesNotExist:
                    return render(request, 'root_app/error-404.html')
                
                except ValueError:
                    return render(request, 'root_app/error-404.html')

                context = {
                    'programme':programme,
                    'is_update':is_update,
                }

                return render(request, 'root_app/programme.html', context)
        
        return render(request, 'root_app/programme.html')

@login_required(login_url='authentication_app:login')
@permission_required('root_app.view_programme', login_url='root_app:permissible_page')
def programmes(request):

    programmes = Programme.objects.all()

    context = {
        'programmes':programmes
    }

    return render(request, 'root_app/programmes.html', context)


@login_required(login_url='authentication_app:login')
@permission_required('root_app.delete_programme', login_url='root_app:permissible_page')
def delete_programme(request):

    try:

        pro_id = request.POST['programme_id']

        Programme.objects.get(id=pro_id).delete()

        return JsonResponse({'code':200, 'message':'Programme has been deleted'})

    except Programme.DoesNotExist:
        return JsonResponse({'code':400, 'message':'Invalid programme id or does not exist'})
    
    except Exception:
        return JsonResponse({'code':400, 'message':'Something went wrong, try again'})


@login_required(login_url='authentication_app:login')
@permission_required('root_app.add_yearclass', login_url='root_app:permissible_page')
def year_class(request, class_id=None):

    is_update = False

    if class_id is not None:
        is_update = True

    if request.method == 'POST':

        class_name = request.POST['class_name']
        pro_id = request.POST['programme_id']
        year = request.POST['year']

        if class_name == '':
            return JsonResponse({'code':400, 'message':'Class name required'})
        
        if pro_id == '':
            return JsonResponse({'code':400, 'message':'Programme required'})

        if is_update:

            if YearClass.objects.filter(class_name=class_name).exists() and YearClass.objects.get(id=class_id).class_name != class_name:
                return JsonResponse({'code':400, 'message':'Class name already exist'})

            year_class = YearClass.objects.get(id=class_id)
            year_class.class_name=class_name
            year_class.year=year
            year_class.programme=Programme.objects.get(id=pro_id)
            year_class.save()
        
            return JsonResponse({'code':200, 'message':'Class updated successfully'})
        
        else:

            if YearClass.objects.filter(class_name=class_name).exists():
                return JsonResponse({'code':400, 'message':'Class name already exist'})

            year_class = YearClass.objects.create(
                class_name=class_name,
                year=year,
                programme=Programme.objects.get(id=pro_id)
            )

            year_class.save()

            return JsonResponse({'code':200, 'message':'Class added successfully'})

    else:

        programmes = Programme.objects.all()

        context = {
            'programmes':programmes
        }

        if is_update:

            try:
                year_class = YearClass.objects.get(id=class_id)
            except YearClass.DoesNotExist:
                    return render(request, 'root_app/error-404.html')
            except ValueError:
                return render(request, 'root_app/error-404.html')

            base_cont = {
                'class':year_class,
                'is_update':is_update,
            }

            context.update(base_cont)

            if request.user.has_perm('root_app.change_yearclass'):
                return render(request, 'root_app/class.html', context)
                
            else:
                return redirect('root_app:permissible_page')
        
        return render(request, 'root_app/class.html', context)


@login_required(login_url='authentication_app:login')
@permission_required('root_app.view_yearclass', login_url='root_app:permissible_page')
def classes(request):

    classes = YearClass.objects.all()

    context = {
        'classes':classes
    }

    return render(request, 'root_app/classes.html', context)


@login_required(login_url='authentication_app:login')
@permission_required('root_app.delete_yearclass', login_url='root_app:permissible_page')
def delete_class(request):

    try:
        class_id = request.POST['class_id']

        year_class = YearClass.objects.get(id=class_id)
        year_class.delete()

        return JsonResponse({'code':200, 'message':'Class deleted successfully'})
    
    except Programme.DoesNotExist:
        return JsonResponse({'code':400, 'message':'Invalid class id or does not exist'})
    
    except Exception:
        return JsonResponse({'code':400, 'message':'Something went wrong, try again'})


@login_required(login_url='authentication_app:login')
@permission_required('root_app.add_hall', login_url='root_app:permissible_page')
def hall(request, hall_id=None):
    is_update = False

    if hall_id is not None:
        is_update = True

    if request.method == 'POST':
        hall_name = request.POST['hall_name']

        if hall_name == '':
            return JsonResponse({'code':400, 'message':'Hall name required'})

        if is_update:
            hall = Hall.objects.get(id=hall_id)
            hall.hall_name=hall_name
            hall.save()

            return JsonResponse({'code':200, 'message':'Hall name updated successfully'})

        else:
            hall = Hall.objects.create(
                hall_name=hall_name
            )
            hall.save()
            return JsonResponse({'code':200, 'message':'Hall added successfully'})
    else:

        if is_update:

            try:
                hall = Hall.objects.get(id=hall_id)

            except Hall.DoesNotExist:
                return render(request, 'root_app/error-404.html')
            
            except ValueError:
                return render(request, 'root_app/error-404.html')

            context = {
                'hall':hall,
                'is_update':is_update
            }

            if not request.user.has_perm('root_app.change_hall'):
                return redirect('root_app:permissible_page')
            
            return render(request, 'root_app/hall.html', context)
        
        return render(request, 'root_app/hall.html')


@login_required(login_url='authentication_app:login')
@permission_required('root_app.view_hall', login_url='root_app:permissible_page')
def halls(request):

    halls = Hall.objects.all()

    context = {
        'halls':halls
    }

    return render(request, 'root_app/halls.html', context)

@login_required(login_url='authentication_app:login')
@permission_required('root_app.delete_hall', login_url='root_app:permissible_page')
def delete_hall(request):

    try:
        hall_id = request.POST['hall_id']

        year_class = Hall.objects.get(id=hall_id)
        year_class.delete()

        return JsonResponse({'code':200, 'message':'Hall deleted successfully'})
    
    except Programme.DoesNotExist:
        return JsonResponse({'code':400, 'message':'Invalid halls id or does not exist'})
    
    except Exception:
        return JsonResponse({'code':400, 'message':'Something went wrong, try again'})


@login_required(login_url='authentication_app:login')
@permission_required('root_app.add_pollingstation', login_url='root_app:permissible_page')
def polling_station(request, station_id=None):

    is_update = False

    if station_id is not None:
        is_update = True
    
    if request.method == 'POST':
        name = request.POST['name']
        location = request.POST['location']

        if name == '':
            return JsonResponse({'code':400, 'message':'Polling station name required'})

        if is_update:
            polling_station = PollingStation.objects.get(id=station_id)
            polling_station.name=name
            polling_station.location=location
            polling_station.save()
            return JsonResponse({'code':200, 'message':'Polling station updated successfuuly'})
        else:
            
            polling_station = PollingStation.objects.create(
                name=name,
                location=location
            )
            polling_station.save()
            return JsonResponse({'code':200, 'message':'Polling station added successfuuly'})

    else:

        if is_update:
            try: 
                polling_station = PollingStation.objects.get(id=station_id)
            
            except PollingStation.DoesNotExist:
                return render(request, 'root_app/error-404.html')
            
            except ValueError:
                return render(request, 'root_app/error-404.html')
            
            context = {
                'polling_station':polling_station,
                'is_update':is_update,
            }

            if not request.user.has_perm('root_app.change_pollingstation'):
                return redirect('root_app:permissible_page')

            return render(request, 'root_app/polling_station.html', context)
        
        else:

            return render(request, 'root_app/polling_station.html')


@login_required(login_url='authentication_app:login')
@permission_required('root_app.view_pollingstation', login_url='root_app:permissible_page')
def polling_stations(request):

    polling_stations = PollingStation.objects.all()

    context = {
        'polling_stations':polling_stations
    }

    return render(request, 'root_app/polling_stations.html', context)



@login_required(login_url='authentication_app:login')
@permission_required('root_app.delete_pollingstation', login_url='root_app:permissible_page')
def delete_pollingstation(request):

    try:
        poll_id = request.POST['poll_id']

        station = PollingStation.objects.get(id=poll_id)
        station.delete()

        return JsonResponse({'code':200, 'message':'Polling station deleted successfully'})
    
    except Programme.DoesNotExist:
        return JsonResponse({'code':400, 'message':'Invalid pollation station id or does not exist'})
    
    except Exception:
        return JsonResponse({'code':400, 'message':'Something went wrong, try again'})


@login_required(login_url='authentication_app:login')
@permission_required('root_app.add_election', login_url='root_app:permissible_page')
def election(request):

    if request.method == 'POST':

        election_name = request.POST['election_name']
        election_year = request.POST['election_year']
        election_date = request.POST['election_date']
        election_time = request.POST['election_time']
        ending_date = request.POST['ending_date']
        ending_time = request.POST['ending_time']

        if request.user.has_perm('root_app.can_assign_commnissioner_role'):
            commissioner = request.POST['electoral_commissioner']
        else:
            commissioner = request.user.id

        polling_stations = request.POST.getlist('polling_stations')

        if election_name == '':
            return JsonResponse({'code':400, 'message':'Election name required'})
        
        elif election_year == '':
            return JsonResponse({'code':400, 'message':'Election year required'})
        
        elif election_date == '':
            return JsonResponse({'code':400, 'message':'Election date required'})

        elif election_time == '':
            return JsonResponse({'code':400, 'message':'Election time required'})
        
        elif ending_date == '':
            return JsonResponse({'code':400, 'message':'Ending date required'})
        
        elif ending_time == '':
            return JsonResponse({'code':400, 'message':'Ending time required'})
        
        elif commissioner == '':
            return JsonResponse({'code':400, 'message':'Choose electoral commissioner'})
        
        if len(polling_stations) <= 0:
            return JsonResponse({'code':400, 'message':'Choose polling station for this election'})


        election = Election.objects.create(
            election_name=election_name,
            election_year=election_year,
            election_date=election_date,
            election_time=election_time,
            ending_date=ending_date,
            ending_time=ending_time,
        )
        
        try:
            request.POST['status']

        except KeyError:
            election.is_active=False
            election.save()

        if request.POST['mul_votes'] == str(1):
            election.allow_multiple_votes=True
            election.save()
        
        else:
            pass
        
        if commissioner != '':

            ElectoralCommissioner.objects.create(
                election=election,
                user=User.objects.get(id=commissioner)
            ).save()
                
        for data in polling_stations:
            AllowedPollingStation.objects.create(
                election=election,
                polling_station=PollingStation.objects.get(id=data)
            ).save()
        
        return JsonResponse({'code':200, 'message':'Election added successfully'})

    else:

        polling_stations = PollingStation.objects.all()
        users = User.objects.all()

        context = {
            'polling_stations':polling_stations,
            'users':users
        }

        return render(request, 'root_app/election.html', context)


@login_required(login_url='authentication_app:login')
@permission_required('root_app.view_election', login_url='root_app:permissible_page')
def elections(request):

    elections = Election.objects.all()     

    context = {
        'elections':elections
    }

    return render(request, 'root_app/elections.html', context)


@login_required(login_url='authentication_app:login')
@permission_required('root_app.change_election', login_url='root_app:permissible_page')
def edit_election(request, election_id):

    if request.method == 'POST':

        election = Election.objects.get(id=election_id)

        election_name = request.POST['election_name']
        election_year = request.POST['election_year']
        election_date = request.POST['election_date']
        election_time = request.POST['election_time']
        ending_date = request.POST['ending_date']
        ending_time = request.POST['ending_time']

        if request.user.has_perm('root_app.can_assign_commnissioner_role'):
            commissioner = request.POST['electoral_commissioner']
        else:
            commissioner = request.user.id

        polling_stations = request.POST.getlist('polling_stations')

        if election_name == '':
            return JsonResponse({'code':400, 'message':'Election name required'})
        
        elif election_year == '':
            return JsonResponse({'code':400, 'message':'Election year required'})
        
        elif election_date == '':
            return JsonResponse({'code':400, 'message':'Election date required'})

        elif election_time == '':
            return JsonResponse({'code':400, 'message':'Election time required'})
        
        elif ending_date == '':
            return JsonResponse({'code':400, 'message':'Ending date required'})
        
        elif ending_time == '':
            return JsonResponse({'code':400, 'message':'Ending time required'})
        
        elif commissioner == '':
            return JsonResponse({'code':400, 'message':'Choose electoral commissioner'})
        
        if len(polling_stations) <= 0:
            return JsonResponse({'code':400, 'message':'Choose polling station for this election'})

        election.election_name=election_name
        election.election_year=election_year
        election.election_date=election_date
        election.election_time=election_time
        election.ending_date=ending_date
        election.ending_time=ending_time
        
        
        try:
            request.POST['status']

            election.is_active=True
            election.save()

        except KeyError:
            election.is_active=False
            election.save()

        if str(request.POST['mul_votes']) == str(1):
            election.allow_multiple_votes=True
            election.save()
        
        else:
            election.allow_multiple_votes=False
            election.save()

        
        if commissioner != '':

            commis = ElectoralCommissioner.objects.get(election=election)
            commis.user=User.objects.get(id=commissioner)
            commis.save()
        
        for data in AllowedPollingStation.objects.filter(election=election):
            data.delete()
                
        for data in polling_stations:
            AllowedPollingStation.objects.create(
                election=election,
                polling_station=PollingStation.objects.get(id=data)
            ).save()
        
        return JsonResponse({'code':200, 'message':'Election updated successfully'})
    else:
        try:
            election = Election.objects.get(id=election_id)

            poll_sta = PollingStation.objects.all()

            users = User.objects.all()

            polling_stations = []

            for data in poll_sta:
                if AllowedPollingStation.objects.filter(election=election, polling_station=PollingStation.objects.get(id=data.id)).exists():
                    polling_stations.append(
                        {
                            'id': data.id,
                            'name': data.name,
                            'poll_exist':data.is_alreadyallowed,
                            'exist':True
                        }
                    )
                else:

                    polling_stations.append(
                        {
                            'id': data.id,
                            'name': data.name,
                            'poll_exist':data.is_alreadyallowed,
                            'exist':False
                        }
                    )

            context = {
                'election':election,
                'polling_stations':polling_stations,
                'users':users
            }

            return render(request, 'root_app/edit_election.html', context)
        
        except Election.DoesNotExist:
            return render(request, 'root_app/error-404.html')
        
        except ValueError:
            return render(request, 'root_app/error-404.html')


@login_required(login_url='authentication_app:login')
@permission_required('root_app.delete_election', login_url='root_app:permissible_page')
def delete_election(request):

    try:
        elec_id = request.POST['elec_id']

        election = Election.objects.get(id=elec_id)
        election.delete()

        return JsonResponse({'code':200, 'message':'Election deleted successfully'})
    
    except Election.DoesNotExist:
        return JsonResponse({'code':400, 'message':'Election id or does not exist'})
    
    except Exception:
        return JsonResponse({'code':400, 'message':'Something went wrong, try again'})



@login_required(login_url='authentication_app:login')
@permission_required('root_app.add_electorate', login_url='root_app:permissible_page')
def add_electorate(request):

    if request.method == 'POST':
        
        index_number = request.POST['index_number']
        email = request.POST['email']
        first_name = request.POST['first_name']
        last_name = request.POST['last_name']
        othername = request.POST['othername']
        class_id = request.POST['class']
        hall_id = request.POST['hall']
        polling_station_id = request.POST['polling_station']
        phone = request.POST['phone']

        if index_number == '':
            return JsonResponse({'code':400, 'message':'Index number is required'})
        
        elif email == '':
            return JsonResponse({'code':400, 'message':'Email is required'})
        
        elif first_name == '':
            return JsonResponse({'code':400, 'message':'First name is required'})
        
        elif last_name == '':
            return JsonResponse({'code':400, 'message':'Last name is required'})
        
        elif phone == '':
            return JsonResponse({'code':400, 'message':'Phone number is required'})
        
        elif User.objects.filter(phone_number=phone).exists():
            return JsonResponse({'code':400, 'message':'Phone number already exist'})
        
        elif User.objects.filter(username=index_number).exists():
            return JsonResponse({'code':400, 'message':'Username already exist'})
        
        elif User.objects.filter(email=email).exists():
            return JsonResponse({'code':400, 'message':'Email already exist'})
        
        elif class_id == '':
            return JsonResponse({'code':400, 'message':'Class is required'})
        
        elif hall_id == '':
            return JsonResponse({'code':400, 'message':'Hall is required'})
        
        elif polling_station_id == '':
            return JsonResponse({'code':400, 'message':'Polling station is required'})
        
        else:

            user = User.objects.create_user(
                username=index_number,
                email=email,
                first_name=first_name,
                last_name=last_name,
                other_name=othername,
                phone_number=phone,
                password=generate_password(10)
            )

            ElectorateProfile.objects.create(
                user=user,
                year_class=YearClass.objects.get(id=class_id),
                hall=Hall.objects.get(id=hall_id),
                polling_station=PollingStation.objects.get(id=polling_station_id)
            )

            return JsonResponse({'code':200, 'message':'Electorate added successfully'})

    else:

        programmes = Programme.objects.all()

        halls = Hall.objects.all()

        polling_stations = PollingStation.objects.all()

        context = {
            'programmes':programmes,
            'halls':halls,
            'polling_stations':polling_stations
        }

        return render(request, 'root_app/electorate.html', context)


@login_required(login_url='authentication_app:login')
def get_class(request):

    try:

        programme_id = request.POST['programme_id']

        programme = Programme.objects.get(id=programme_id)

        year_class = YearClass.objects.filter(programme=programme)

        return JsonResponse({'class':list(year_class.values())})
    
    except Exception:

        return JsonResponse({'class':[]})
    


@login_required(login_url='authentication_app:login')
@permission_required('root_app.add_electorate', login_url='root_app:permissible_page')
def electorates(request):

    users = User.objects.all()

    electorates = []

    for data in users:
        if data.user_type == 'user':
            electorates.append(
                {
                    'id':data.id,
                    'index_number':data.username,
                    'full_name':data.full_name,
                    'class': data.electorateprofile.year_class.class_name,
                    'polling_station':data.electorateprofile.polling_station.name
                }
            )

    context = {
        'electorates':electorates
    }

    return render(request, 'root_app/electorates.html', context)


@login_required(login_url='authentication_app:login')
@permission_required('root_app.change_electorate', login_url='root_app:permissible_page')
def edit_electorate(request, electorate_id):

    if request.method == 'POST':

        user = User.objects.get(id=electorate_id)
        
        index_number = request.POST['index_number']
        email = request.POST['email']
        first_name = request.POST['first_name']
        last_name = request.POST['last_name']
        othername = request.POST['othername']
        class_id = request.POST['class']
        hall_id = request.POST['hall']
        polling_station_id = request.POST['polling_station']
        phone = request.POST['phone']

        if index_number == '':
            return JsonResponse({'code':400, 'message':'Index number is required'})
        
        elif email == '':
            return JsonResponse({'code':400, 'message':'Email is required'})
        
        elif first_name == '':
            return JsonResponse({'code':400, 'message':'First name is required'})
        
        elif last_name == '':
            return JsonResponse({'code':400, 'message':'Last name is required'})
        
        elif phone == '':
            return JsonResponse({'code':400, 'message':'Phone number is required'})
        
        elif User.objects.filter(phone_number=phone).exists() and user.phone_number != phone:
            return JsonResponse({'code':400, 'message':'Phone number already exist'})
        
        elif User.objects.filter(username=index_number).exists() and user.username != index_number:
            return JsonResponse({'code':400, 'message':'Username already exist'})
        
        elif User.objects.filter(email=email).exists() and user.username != index_number:
            return JsonResponse({'code':400, 'message':'Email already exist'})
        
        elif class_id == '':
            return JsonResponse({'code':400, 'message':'Class is required'})
        
        elif hall_id == '':
            return JsonResponse({'code':400, 'message':'Hall is required'})
        
        elif polling_station_id == '':
            return JsonResponse({'code':400, 'message':'Polling station is required'})
        
        else:

            user.username=index_number
            user.email=email
            user.first_name=first_name
            user.last_name=last_name
            user.other_name=othername
            user.phone_number=phone
            user.password=generate_password(10)
            user.save()

            electorate = ElectorateProfile.objects.get(id=user.electorateprofile.id)

            electorate.year_class=YearClass.objects.get(id=class_id)
            electorate.hall=Hall.objects.get(id=hall_id)
            electorate.polling_station=PollingStation.objects.get(id=polling_station_id)
            electorate.save()

            return JsonResponse({'code':200, 'message':'Electorate has been updated successfully'})

    else:
        electorate = User.objects.get(id=electorate_id)

        programmes = Programme.objects.all()

        halls = Hall.objects.all()

        polling_stations = PollingStation.objects.all()

        context = {
            'programmes':programmes,
            'halls':halls,
            'polling_stations':polling_stations,
            'electorate':electorate,
        }

        return render(request, 'root_app/edit_electorate.html', context)


@login_required(login_url='authentication_app:login')
@permission_required('root_app.upload_electorate_with_excel', login_url='root_app:permissible_page')
def electorate_excelupload(request):

    if request.method == 'POST':

        try:
            excel_file = request.FILES['excel_file']
        
        except KeyError:
            return JsonResponse({'code':400, 'message':'No file selected'})

        class_id = request.POST['year_class']

        hall_id = request.POST['hall']

        polling_station_id = request.POST['polling_station']

        
        if class_id == '':
            return JsonResponse({'code':400, 'message':'No class is selected'})
        
        elif hall_id ==  '':
            return JsonResponse({'code':400, 'message':'No hall is selected'})
        
        elif polling_station_id == '':
            return JsonResponse({'code':400, 'message':'No polling station is selected'})
        
        book = load_workbook(excel_file)
        
        sheet = book.active

        rows = sheet.rows

        headers = [data.value for data in next(rows)]

        all_data = []

        for row in rows:

            data = {}

            for key, value in zip(headers, row):
                data[key] = value.value
            
            all_data.append(data)
        
        data_response = []


        try:


            for data in all_data:

                if data['index_number'] is None:

                    data['response_message'] = 'Index number is required'
                    
                    data_response.append(data)

                elif User.objects.filter(username=data['index_number']).exists():

                    data['response_message'] = 'Index number already taken'

                    data_response.append(data)

                elif data['first_name'] is None:

                    data['response_message'] = 'First name is required'

                    data_response.append(data)
            
                elif data['last_name'] is None:

                    data['response_message'] = 'Last name is required'

                    data_response.append(data)
                
                elif data['email'] is None:

                    data['response_message'] = 'Email is required'

                    data_response.append(data)

                elif User.objects.filter(phone_number=data['phone_number']).exists():

                    data['response_message'] = 'Phone number is already taken'

                    data_response.append(data)

                elif data['phone_number'] is None:

                    data['response_message'] = 'Phone number is required'

                    data_response.append(data)
                
                elif User.objects.filter(email=data['email']).filter():

                    data['response_message'] = 'Email already taken'

                    data_response.append(data)
                
                else:

                    user = User.objects.create_user(
                        username=data['index_number'],
                        first_name=data['first_name'],
                        last_name=data['last_name'],
                        other_name=data['other_name'],
                        email=data['email'],
                        phone_number=data['phone_number'],
                        password=generate_password(10)
                    )

                    ElectorateProfile.objects.create(
                        user=user,
                        year_class=YearClass.objects.get(id=class_id),
                        hall=Hall.objects.get(id=hall_id),
                        polling_station=PollingStation.objects.get(id=polling_station_id)
                    )
            
            return JsonResponse({'code':200, 'message':'All data uploaded successfully', 'response_data':data_response, 'response_length':len(data_response), 'data_report':f'{len(data_response)} out {len(all_data)} data was unable to upload. Check from table to see those data with errors.'})
        
        except Exception as e:
            return JsonResponse({'code':500, 'message':str(e)})
    else:

        programmes = Programme.objects.all()

        halls = Hall.objects.all()

        polling_stations = PollingStation.objects.all()

        context = {
            'programmes':programmes,
            'halls':halls,
            'polling_stations':polling_stations,

        }

        return render(request, 'root_app/electorate_excel_upload.html', context)

def get_excel_data(request):

    try:

        excel_file = request.FILES['excel_file']

        book = load_workbook(excel_file)

        sheet = book.active

        rows = sheet.rows

        headers = [ data.value for data in next(rows) ]

        all_data =[]

        for row in rows:

            data_list = []

            for key, value in zip(headers, row):
                data_list.append(value.value)
            all_data.append(data_list)

        return JsonResponse({'headers':headers, 'all_data':all_data})
    
    except Exception:
        return JsonResponse({'headers':[], 'all_data':[]})

@login_required(login_url='authentication_app:login')
@permission_required('authentication_app.delete_user', login_url='root_app:permissible_page')
def delete_electorate(request):

    try:
        user_id = request.POST['user_id']

        electrorate = User.objects.get(id=user_id)
        electrorate.delete()

        return JsonResponse({'code':200, 'message':'Electorate deleted successfully'})
    
    except User.DoesNotExist:
        return JsonResponse({'code':400, 'message':'Electorate id or does not exist'})
    
    except Exception:
        return JsonResponse({'code':400, 'message':'Something went wrong, try again'})


@login_required(login_url='authentication_app:login')
@permission_required('authentication_app.delete_user', login_url='root_app:permissible_page')
def delete_bulkelectorates(request):

    ids = request.POST.getlist('ids[]')

    for data in ids:
        User.objects.get(id=data).delete()

    return JsonResponse({'code':200, 'message':'Electorate(s) deleted successfully'})


@login_required(login_url='authentication_app:login')
@permission_required('root_app.add_position', login_url='root_app:permissible_page')
def position(request, position_id=None):

    is_update = False

    if position_id is not None:
        is_update = True

    if request.method == 'POST':
        election_id = request.POST['election']
        position_name = request.POST['position_name']
        description = request.POST['description']
        number_of_asp = request.POST['no_aspirants']

        if election_id == '':
            return JsonResponse({'code':400, 'message':'Select Election'})
        
        elif position_name  == '':
            return JsonResponse({'code':400, 'message':'Position name is required'})

        elif number_of_asp  == '':
            return JsonResponse({'code':400, 'message':'Number of aspirants required'})

        if is_update:

            position = Position.objects.get(id=position_id)
            position.election=Election.objects.get(id=election_id)
            position.position_name=position_name
            position.position_description=description
            position.number_of_asp=number_of_asp
            position.save()

        else:   

            Position.objects.create(
                election=Election.objects.get(id=election_id),
                position_name=position_name,
                number_of_asp=number_of_asp,
                position_description=description
            )
        
        if is_update:
            return JsonResponse({'code':200, 'message':'Position is updated successfully'})
        else:
            return JsonResponse({'code':200, 'message': 'New position is added successfully'})
    else:

        elections = Election.objects.all()     

        context = {
            'elections':elections
        }

        if is_update:
            
            try:

                position = Position.objects.get(id=position_id)

                new_context = {
                    'position':position,
                    'is_update':is_update,
                }

                context.update(new_context)

                return render(request, 'root_app/position.html', context)
            
            except Position.DoesNotExist:
                return render(request, 'root_app/error-404.html')
            
            except ValueError:
                return render(request, 'root_app/error-404.html')

        else:
            return render(request, 'root_app/position.html', context)
        


@login_required(login_url='authentication_app:login')
@permission_required('root_app.view_position', login_url='root_app:permissible_page')
def positions(request):

    posts = Position.objects.all()

    elections = Election.objects.all()

    context = {
        'positions':posts,
        'elections':elections,
    }

    return render(request, 'root_app/positions.html', context)


@login_required(login_url='authentication_app:login')
@permission_required('authentication_app.delete_position', login_url='root_app:permissible_page')
def delete_position(request):

    try:
        position_id = request.POST['position_id']

        position = Position.objects.get(id=position_id)
        position.delete()

        return JsonResponse({'code':200, 'message':'Position deleted successfully'})
    
    except User.DoesNotExist:
        return JsonResponse({'code':400, 'message':'Position id or does not exist'})
    
    except Exception:
        return JsonResponse({'code':400, 'message':'Something went wrong, try again'})



def position_filter(request):

    try:
        election_id = request.POST['election']

        election = Position.objects.filter(election_id=Election.objects.get(id=election_id))
    
    except ValueError:
        if request.user.has_perm('root_app.add_election') and request.user.user_type != 'superuser' and not request.user.has_perm('root_app.can_assign_commnissioner_role'):
            print(True)
            return JsonResponse({'data':[]})
        
        election = Position.objects.all()

    positions = []

    for data in election:

        post = {
                'id':data.id,
                'position_name':data.position_name,
                'election':data.election.election_name,
                'url': reverse('root_app:position', args=[data.id]),
                'can_delete':False,
                'can_change':False,
            }

        positions.append(post)

        if request.user.has_perm('root_app.delete_position'):
            post.update({'can_delete':True})
        
        if request.user.has_perm('root_app.change_position'):
            post.update({'can_change':True})


    return JsonResponse({'data':positions})


@login_required(login_url='authentication_app:login')
@permission_required('root_app.add_aspirant', login_url='root_app:permissible_page')
def add_aspirant(request):

    if request.method == "POST":
        
        first_name = request.POST['first_name']
        surname = request.POST['surname']
        other_name =request.POST['othername']
        election_id = request.POST['election']
        position_id = request.POST['position']
        ballot_number = request.POST['ballot_number']

        try:
            image = request.FILES['passport_pic']
        
        except Exception:
            return JsonResponse({'code':400, 'message':'Select picture for candidate'})
        
        if first_name == '':
            return JsonResponse({'code':400, 'message':'First name is required'})
        
        elif surname == '':
            return JsonResponse({'code':400, 'message':'Surname is required'})
        
        elif election_id == '':
            return JsonResponse({'code':400, 'message':'Choose a specific election for this cadidate'})
        
        elif position_id == '':
            return JsonResponse({'code':400, 'message':'Choose a position for this candidate'})
        
        elif ballot_number == '':
            return JsonResponse({'code':400, 'message':'Select a ballot number for candidate'})

        if Aspirant.objects.filter(position=Position.objects.get(id=position_id), ballot_number=ballot_number).exists():
            return JsonResponse({'code':400, 'message':'Ballot number for the selected position is already taken'})


        
        Aspirant.objects.create(
            passport_picture=image,
            first_name=first_name,
            surname=surname,
            other_name=other_name,
            election=Election.objects.get(id=election_id),
            position=Position.objects.get(id=position_id),
            ballot_number=ballot_number
        )

        return JsonResponse({"code":200, 'message':'New aspirant added successfully'})

    else:

        elections = Election.objects.all()


        context = {
            'elections':elections,
        }

        return render(request, 'root_app/add_aspirant.html', context)


@login_required(login_url='authentication_app:login')
def filter_ballotnumber(request):

    try:

        post_id = request.POST['position']

        position = Position.objects.get(id=post_id)

        asp = Aspirant.objects.filter(position=position)

        numbers = [data.ballot_number for data in asp]

        available_number = []

        for posts in range(1, position.number_of_asp + 1):
            if posts in numbers:
                continue
            else:
                available_number.append(posts)

        return JsonResponse({'data':available_number})

    except Exception:
        return JsonResponse({'data':[]})



@login_required(login_url='authentication_app:login')
@permission_required('root_app.can_view_aspirants', login_url='root_app:permissible_page')
def aspirants(request):

    aspirants = Aspirant.objects.all()

    elections = Election.objects.all()

    context = {
        'aspirants':aspirants,
        'elections':elections
    }

    return render(request, 'root_app/aspirants.html', context)




@login_required(login_url='authentication_app:login')
@permission_required('root_app.change_aspirant', login_url='root_app:permissible_page')
def edit_aspirant(request, aspirant_id):

    if request.method == "POST":
        
        first_name = request.POST['first_name']
        surname = request.POST['surname']
        other_name =request.POST['othername']
        election_id = request.POST['election']
        position_id = request.POST['position']
        ballot_number = request.POST['ballot_number']
        old_image = request.POST['old_image']

        try:
            image = request.FILES['passport_pic']
        
        except Exception:
            image = old_image

        aspirant = Aspirant.objects.get(id=aspirant_id)
        
        if first_name == '':
            return JsonResponse({'code':400, 'message':'First name is required'})
        
        elif surname == '':
            return JsonResponse({'code':400, 'message':'Surname is required'})
        
        elif election_id == '':
            return JsonResponse({'code':400, 'message':'Choose a specific election for this cadidate'})
        
        elif position_id == '':
            return JsonResponse({'code':400, 'message':'Choose a position for this candidate'})
        
        elif ballot_number == '':
            return JsonResponse({'code':400, 'message':'Select a ballot number for candidate'})

        elif Aspirant.objects.filter(position=Position.objects.get(id=position_id), ballot_number=ballot_number).exists() and aspirant.ballot_number != int(ballot_number):
            return JsonResponse({'code':400, 'message':'Ballot number for the selected position is already taken'})

        if ballot_number == '':
            return JsonResponse({'code':400, 'message':'Select a ballot number for candidate'})

        
        aspirant.passport_picture=image
        aspirant.first_name=first_name
        aspirant.surname=surname
        aspirant.other_name=other_name
        aspirant.election=Election.objects.get(id=election_id)
        aspirant.position=Position.objects.get(id=position_id)
        aspirant.ballot_number=ballot_number
        aspirant.save()

        return JsonResponse({"code":200, 'message':'Aspirant updated successfully'})

    else:

        elections = Election.objects.all()

        try:

            aspirant = Aspirant.objects.get(id=aspirant_id)
        
        except Aspirant.DoesNotExist:
            return render(request, 'root_app/error-404.html')
        
        except ValueError:
            return render(request, 'root_app/error-404.html')


        asp = Aspirant.objects.filter(position=aspirant.position)

        numbers = [data.ballot_number for data in asp]

        available_number = []

        for posts in range(1, aspirant.position.number_of_asp + 1):
            if posts in numbers:
                continue
            else:
                available_number.append(posts)

        context = {
            'elections':elections,
            'aspirant':aspirant,
            'ballot_numbers':available_number
        }

        return render(request, 'root_app/edit_aspirant.html', context)



@login_required(login_url='authentication_app:login')
@permission_required('root_app.delete_aspirant', login_url='root_app:permissible_page')
def delete_aspirant(request):
    try:
        aspirant_id = request.POST['aspirant_id']

        aspirant = Aspirant.objects.get(id=aspirant_id)
        aspirant.delete()

        return JsonResponse({'code':200, 'message':'Aspirant deleted successfully'})
    
    except User.DoesNotExist:
        return JsonResponse({'code':400, 'message':'Aspirant id or does not exist'})
    
    except Exception:
        return JsonResponse({'code':400, 'message':'Something went wrong, try again'})



@login_required(login_url='authentication_app:login')
@permission_required('root_app.view_aspirant', login_url='root_app:permissible_page')
def view_aspirant(request):

    try:
        aspirant_id = request.POST['aspirant_id']

        aspirant = Aspirant.objects.get(id=aspirant_id)

        context = {
            'photo': aspirant.passport_picture.url,
            'full_name':aspirant.full_name,
            'election':aspirant.election.election_name,
            'position':aspirant.position.position_name,
            'ballot_number':aspirant.ballot_number,
            'date_added':f'{aspirant.time_stamp.date()} at {aspirant.time_stamp.time()}',
        }

        return JsonResponse(context)
    
    except Exception:
        return JsonResponse({})


@login_required(login_url='authentication_app:login')
def filter_aspirant(request):

    election_id = request.POST['election']
    position_id = request.POST['position']

    try:

        if election_id != '' and  position_id == '':

            aspirants = Aspirant.objects.filter(election=Election.objects.get(id=election_id))
        
        elif position_id != '':

            aspirants = Aspirant.objects.filter(position=Position.objects.get(id=position_id))

        else:

            aspirants = Aspirant.objects.all()

        filtered_asps = []
        
        if request.user.has_perm('root_app.add_election') and request.user.user_type != 'superuser' and not request.user.has_perm('root_app.can_assign_commnissioner_role'):


            for data in aspirants:

                if data.election.electoralcommissioner.user.id == request.user.id:
                    
                    detail = {
                        'id':data.id,
                        'full_name':data.full_name,
                        'election': data.election.election_name,
                        'position':data.position.position_name,
                        'ballot_number':data.ballot_number,
                        'photo':data.passport_picture.url,
                        'url': reverse('root_app:edit_aspirant', args=[data.id]),
                        'can_edit':False,
                        'can_view':False,
                        'can_delete':False
                    }


                    if request.user.has_perm('root_app.change_aspirant'):
                        detail['can_edit'] = True

                    if request.user.has_perm('root_app.view_aspirant'):
                        detail['can_view'] = True

                    if request.user.has_perm('root_app.delete_aspirant'):
                        detail['can_delete'] = True 

                    filtered_asps.append(
                        detail
                    ) 
                
        else:

            for data in aspirants:
                detail = {
                    'id':data.id,
                    'full_name':data.full_name,
                    'election': data.election.election_name,
                    'position':data.position.position_name,
                    'ballot_number':data.ballot_number,
                    'photo':data.passport_picture.url,
                    'url': reverse('root_app:edit_aspirant', args=[data.id]),
                    'can_edit':False,
                    'can_view':False,
                    'can_delete':False
                }
                
                if request.user.has_perm('root_app.change_aspirant'):
                    detail['can_edit'] = True

                if request.user.has_perm('root_app.view_aspirant'):
                    detail['can_view'] = True

                if request.user.has_perm('root_app.delete_aspirant'):
                    detail['can_delete'] = True
                
                filtered_asps.append(
                    detail
                )

        return JsonResponse({'data':filtered_asps})
    
    except Aspirant.DoesNotExist:

        return JsonResponse({})


@login_required(login_url='authentication_app:login')
@permission_required('root_app.view_election_results_list', login_url='root_app:permissible_page')
def election_results(request):

    elections = Election.objects.all()     

    context = {
        'elections':elections
    }

    return render(request, 'root_app/election_results.html', context)


@login_required(login_url='authentication_app:login')
@permission_required('root_app.verify_electorate', login_url='root_app:permissible_page')
def verification(request):

    if request.method == 'POST':
        username = request.POST['username']

        cur_election = CurrentElection.objects.all().last()
        election = Election.objects.get(id=cur_election.election.id)

        if election == '':
            return JsonResponse({'code':400, 'message':'Select Election'})
            
        elif username == '':
            return JsonResponse({'code':400, 'message':'Enter a username to verify'})

        if User.objects.filter(username=username).exists() and User.objects.get(username=username).user_type == 'user':

            if VerifiedElectorate.objects.filter(user=User.objects.get(username=username), election=election).exists():
                return JsonResponse({'code':400, 'message':'User has already been verified'})
            
            password = generate_password(10)

            user = User.objects.get(username=username)
            user.set_password(password)
            user.save()

            VerifiedElectorate.objects.create(
                user=user,
                election=election,
                verified_by=request.user
            )

            context = {
                'full_name':user.full_name,
                'year_class':user.electorateprofile.year_class.class_name,
                'password':password
            }

            return JsonResponse({'code':200, 'message':'Verified', 'user_detail':context})
        else:

            return JsonResponse({'code':400, 'message':'User does not exist or does not exist as a voter'})

    else:
        elections = Election.objects.all()     

        context = {
            'elections':elections
        }

        return render(request, 'root_app/verification_form.html', context)


@login_required(login_url='authentication_app:login')
@permission_required('root_app.view_verifiedelectorates', login_url='root_app:permissible_page')
def verified_electorates(request):

    v_electorates = VerifiedElectorate.objects.all()


    context = {
        'verified_electorates': v_electorates
    }

    return render(request, 'root_app/verified_electorates.html', context)


@login_required(login_url='authentication_app:login')
@permission_required('root_app.unverify_electorates', login_url='root_app:permissible_page')
def unverify_electorate(request):

    try:
        uv_id = request.POST['uv_id']

        VerifiedElectorate.objects.get(id=uv_id).delete()

        return JsonResponse({'code':200, 'message':'Electorate unverified successfully'})
    
    except Exception:

        return JsonResponse({'code':400, 'message':'Something went wrong, please try again'})

@login_required(login_url='authentication_app:login')
@permission_required('root_app.set_currentelection', login_url='root_app:permissible_page')
def set_currentelection(request):

    if request.method == 'POST':

        try:

            election_id = request.POST['election']

            if election_id == '':
                return JsonResponse({'code':400, 'message':'No election selected'})

            election = Election.objects.get(id=election_id)

            CurrentElection.objects.create(
                election=election,
                set_by=request.user
            )

            return JsonResponse({'code':200, 'message':f'{election.election_name} has been set as the current election'})
        
        except Exception as e:
            return JsonResponse({'code':400, 'message':'Something went wrong, try again!'})
        
    else:

        elections = Election.objects.all()     

        context = {
            'elections':elections
        }

        return render(request, 'root_app/current_election.html', context)


@login_required(login_url='authentication_app:login')
def voting_ballot(request):

    if request.method == 'POST':
        pass
    
    else:

        election = CurrentElection.objects.all().last()

        current_election = Election.objects.get(id=election.election.id)

        has_voted = Vote.objects.filter(election=current_election, user=request.user).exists()

        poll_ids = []

        polling_stations = AllowedPollingStation.objects.filter(election=current_election)

        for data in polling_stations:
            poll_ids.append(data.polling_station.id)


        context = {
            'election':current_election,
        }

        if request.user.electorateprofile.polling_station.id in poll_ids:
            context.update(
                {'is_permitted':True}
            )

        if not current_election.allow_multiple_votes and has_voted:
            context.update(
                {'has_voted':True}
            )
            

        return render(request, 'root_app/voting_ballot.html', context)

def get_selected_aspirants(request):

    aspirants_id = request.POST.getlist('asp_ids[]')

    aspirants = []

    for data in aspirants_id:
        asp = Aspirant.objects.get(id=data)
        aspirants.append(
            {
                'full_name':asp.full_name,
                'ballot_number':asp.ballot_number,
                'passport_picture':asp.passport_picture.url,
                'position':asp.position.position_name
            }
        )
    
    return JsonResponse({'data':aspirants})



@login_required(login_url='authentication_app:login')
def cast_vote(request):

    election = CurrentElection.objects.all().last()

    current_election = Election.objects.get(id=election.election.id)

    aspirants_id = request.POST.getlist('asp_ids[]')

    has_voted = Vote.objects.filter(election=current_election, user=request.user).exists()

    if not current_election.allow_multiple_votes and has_voted:
        return JsonResponse({'code':400, 'message':'You have already voted'})
    elif len(aspirants_id) <= 0:
        return JsonResponse({'code':400, 'message':'No candidate selected'})

    for data in aspirants_id:
        Vote.objects.create(
            aspirant=Aspirant.objects.get(id=data),
            user=request.user,
            election=current_election,
            position=Aspirant.objects.get(id=data).position
        )
    
    return JsonResponse({'code':200, 'message':'Your vote submitted successfully', 'url':reverse('authentication_app:logout')})



@login_required(login_url='authentication_app:login')
@permission_required('root_app.view_results', login_url='root_app:permissible_page')
def results(request, election_id):
    try:
        election = Election.objects.get(id=election_id)

        context = {
            'election':election
        }
        
        return render(request, 'root_app/results.html', context)
    
    except Election.DoesNotExist:
        return render(request, 'root_app/error-404.html', )
    
    except ValueError:
        return render(request, 'root_app/error-404.html', )

def get_results(request, election_id):

    election = Election.objects.get(id=election_id)

    results_data = []

    #Getting positions with no results
    for data in election.position_set.all():
        asp = {
            'position_name':data.position_name,
            'results': []
        }

        #Getting results for each position
        for ele in data.aspirant_set.all():
            asp['results'].append(
                {
                    'full_name':ele.full_name,
                    'votes':ele.aspirant_vote_counts,
                    'percentage':ele.aspirant_vote_percentage
                }
            )
        
        results_data.append(asp)

    return JsonResponse({'data':results_data})



@login_required(login_url='authentication_app:login')
@permission_required('root_app.print_results', login_url='root_app:permissible_page')
def print_results(request, election_id):

    try:

        election = Election.objects.get(id=election_id)

        context = {
            'election':election
        }

        return render(request, 'root_app/results_print.html', context)
    
    except Election.DoesNotExist:
        return render(request, 'root_app/error-404.html')
    
    except ValueError:
        return render(request, 'root_app/error-404.html')



@login_required(login_url='authentication_app:login')
@permission_required('root_app.view_detailed_report', login_url='root_app:permissible_page')
def detailed_report(request):

    if request.method == 'POST':
        
        username = request.POST['username']

        if username == '':
            return JsonResponse({'code':400, 'message':'Enter a username'})
        else:
            if not User.objects.filter(username=username).exists():
                return JsonResponse({'code':400, 'message':'Username does not exist'})
            
            elif User.objects.get(username=username).user_type != 'user':
                return JsonResponse({'code':400, 'message':'Cannot pull details for such user'})
            
            else:
                user = User.objects.get(username=username)

                user_details = {
                    'full_name':user.full_name,
                    'email':user.email,
                    'time_stamp':user.time_stamp,
                    'last_login':user.last_login,
                    'programme':user.electorateprofile.year_class.programme.programme_name,
                    'class':user.electorateprofile.year_class.class_name,
                    'hall':user.electorateprofile.hall.hall_name,
                    'polling_station':user.electorateprofile.polling_station.name,
                    'voting_logs':[],
                }

                for data in user.vote_set.all():
                    user_details['voting_logs'].append(
                        {
                            'election':data.election.election_name,
                            'position':data.position.position_name,
                            'time_stamp':data.time_stamp
                        }
                    )
                
                return JsonResponse({'code':200, 'details':user_details})
                
    return render(request, 'root_app/detailed_report.html')


@login_required(login_url='authentication_app:login')
def send_password_link(request):

    try:
        user_id = request.POST['user_id']

        user = User.objects.get(id=user_id)

        if send_password_resetlink(request, user, user.email):
            return JsonResponse({'code':200, 'message':'Email sent successfully'})

        return JsonResponse({'code':400, 'message':'Sending email was unsuccessfull'})
    
    except Exception as e:
        #return JsonResponse({'code':400, 'message':str(e)})
        return JsonResponse({'code':400, 'message':'Something went wrong, try again'})


@login_required(login_url='authentication_app:login')
@permission_required('root_app.view_general_report', login_url='root_app:permissible_page')
def general_report(request):

    group = Group.objects.all()
    user = User.objects.all()
    programmes = Programme.objects.all()
    yearclass = YearClass.objects.all()
    halls = Hall.objects.all()
    polling_station = PollingStation.objects.all()
    elections = Election.objects.all()
    position = Position.objects.all()
    aspirants = Aspirant.objects.all()
    electorates = ElectorateProfile.objects.all()
    verified = VerifiedElectorate.objects.all()

    context = {
        'group':group,
        'users':user,
        'programmes':programmes,
        'yearclass':yearclass,
        'halls':halls,
        'polling_sation':polling_station,
        'elections':elections,
        'positions':position,
        'aspirants':aspirants,
        'electorates':electorates,
        'verified_electorates':verified,
    }

    user_types = {
        'superuser':0,
        'staff': 0,
        'common_user': 0
    }

    verification_bd = {
        'normal':0,
        'self':0,
    }

    for data in user:

        if data.user_type == 'superuser':
            user_types['superuser'] += 1
        elif data.user_type == 'staff':
            user_types['staff'] += 1
        else:
            user_types['common_user'] += 1
    
    for data in verified:
        if data.user == data.verified_by:
            verification_bd['self'] += 1
        else:
            verification_bd['normal'] += 1

    context.update(verification_bd)       
    context.update(user_types)

    return render(request, 'root_app/general_report.html', context)


@login_required(login_url='authentication_app:login')
@permission_required('root_app.view_institutioninfo', login_url='root_app:permissible_page')
def institution_info(request):

    info_id = 1

    if request.method == 'POST':
        name = request.POST['name']
        abv_name = request.POST['abv_name']
        email = request.POST['email']
        phone = request.POST['phone']
        description = request.POST['description']

        if not InstitutionInfo.objects.filter(id=info_id).exists():
            try:
                logo = request.FILES['logo']

            except KeyError:
                return JsonResponse({'code':400, 'message':'Institutions logo required'})

        if name == '':
            return JsonResponse({'code':400, 'message':"Institution's name is required"})
        
        elif abv_name == '':
            return JsonResponse({'code':400, 'message':"Abreviated name is required"})
        
        elif email == '':
            return JsonResponse({'code':400, 'message':"Institution's email is required"})
        
        elif phone == '':
            return JsonResponse({'code':400, 'message':"Institution's phone number is required"})

        try:
            logo = request.FILES['logo']

        except KeyError:

            logo = request.POST['old_logo']

        if InstitutionInfo.objects.filter(id=info_id).exists():
            inst_info = InstitutionInfo.objects.get(id=info_id)
            inst_info.logo=logo
            inst_info.name=name
            inst_info.email=email
            inst_info.phone=phone
            inst_info.description=description
            inst_info.save()

            return JsonResponse({'code':200, 'message':'Institution info updated successfully'})
        else:
            InstitutionInfo.objects.create(
                logo=logo,
                name=name,
                abv_name=abv_name,
                email=email,
                phone=phone,
                description=description
            )

            return JsonResponse({'code':200, 'message':'Institution info created succesfully'})
    
    else:

        try:

            inst_info = InstitutionInfo.objects.get(id=info_id)

            context = {
                'institution':inst_info
            }

            return render(request, 'root_app/institution_info.html', context)
        
        except InstitutionInfo.DoesNotExist:
            return render(request, 'root_app/institution_info.html')